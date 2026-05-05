

import re
import time
import logging
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from urllib.parse import urljoin, urlparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout


# ══════════════════════════════════════════════════════════════════════════════
#  ▶▶▶  EXCEL CONFIG  ◀◀◀
# ══════════════════════════════════════════════════════════════════════════════

MASTER_EXCEL = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\TamilNadu Report.xlsx"
OUTPUT_EXCEL = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\TN.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
#  WEBSITE CONFIG  (no need to change anything below)
# ════════════════════════════════════
BASE_URL = "https://www.tn.gov.in"

# Each category: name shown in Excel col A, URL to scrape, structure flags
CATEGORIES = [
    {
        "name":      "Government order",
        "url":       f"{BASE_URL}/godept_list.php",
        "has_depts": True,
        "has_years": True,
    },
    {
        "name":      "Policy Notes",
        "url":       f"{BASE_URL}/document_dept_list.php?cate_name=YWxs",
        "has_depts": True,
        "has_years": True,
    },
    {
        "name":      "Rules and Regulations",
        "url":       f"{BASE_URL}/rules.php",
        "has_depts": True,
        "has_years": True,
    },
    {
        "name":      "Circulars and Notifications",
        "url":       f"{BASE_URL}/circular_notifications_dept_list.php",
        "has_depts": True,
        "has_years": True,
    },
    {
        "name":      "Acts",
        "url":       f"{BASE_URL}/act_ordinances.php",
        "has_depts": False,   # Acts goes straight to year pages
        "has_years": True,
    },
]

EXCEL_SHEET  = "Document Details"
DELAY        = 1.5       # seconds between downloads
DL_TIMEOUT   = 60        # seconds per HTTP download
PAGE_TIMEOUT = 40_000    # ms Playwright page-load
GENERIC_PDF_NAMES = {
    "document.pdf", "download.pdf", "file.pdf",
    "new.pdf", "unnamed.pdf", "unknown.pdf",
}


# ══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("tn_excel_downloader.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL  —  READ INDEX
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_index(excel_path: str) -> set:
    """
    Read Column C (File Name) from the 'Document Details' sheet.
    Returns a set of lower-cased filenames = PDFs we already have.

    Also stores year-prefix-stripped variants:
      '2024_adtw_e_100.pdf'  →  also stores  'adtw_e_100.pdf'
    so the skip works regardless of naming convention in the Excel.
    """
    log.info("Reading Excel: %s", excel_path)
    try:
        df = pd.read_excel(excel_path, sheet_name=EXCEL_SHEET,
                           header=0, dtype=str)
    except Exception as exc:
        log.error("Cannot open Excel: %s", exc)
        raise

    df.columns = [str(c).strip() for c in df.columns]

    # Only actual PDF data rows — skip TOTAL / header / summary rows
    type_col = "Type" if "Type" in df.columns else df.columns[3]
    data = df[df[type_col].str.strip().str.upper() == "PDF"].copy()

    fname_col = "File Name" if "File Name" in df.columns else df.columns[2]

    index = set()
    for raw in data[fname_col].dropna():
        name = str(raw).strip().lower()
        if not name or name == "nan":
            continue
        index.add(name)
        # strip leading YYYY_ or YYYY- prefix
        stripped = re.sub(r'^\d{4}[-_]', '', name)
        if stripped != name:
            index.add(stripped)

    log.info("  Excel index built: %d known filenames → will all be SKIPPED",
             len(index))
    return index


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL  —  WRITE NEW ROWS
# ══════════════════════════════════════════════════════════════════════════════

def append_to_excel(excel_path: str, new_rows: list, save_path: str = None):
    """
    Insert new PDF records into the Excel file.
    Each row is placed BEFORE its category's TOTAL row (matching your format).
    Falls back to end-of-sheet if TOTAL row not found.

    If save_path is provided, the updated workbook is saved to that path.
    Otherwise the original Excel file is overwritten.
    """
    if not new_rows:
        return

    if save_path is None:
        save_path = excel_path

    log.info("Updating Excel with %d new rows …", len(new_rows))
    wb = load_workbook(excel_path)
    ws = wb[EXCEL_SHEET]

    # ── Styles matching existing PDF rows ───────────────────────────────
    white  = PatternFill("solid", fgColor="FFFFFF")
    red_b  = Font(color="FF0000", bold=True)   # for Type = PDF column
    black  = Font(color="000000")
    ctr    = Alignment(horizontal="center")
    lft    = Alignment(horizontal="left")

    def write_row(row_num, r):
        data = [r["category"], r["sub_category"], r["file_name"],
                r["type"],     r["pdf_count"],    r["word_count"]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = val
            cell.fill  = white
            if col == 4:                      # Type → red bold centred
                cell.font = red_b; cell.alignment = ctr
            elif col in (5, 6):              # counts → centred
                cell.font = black; cell.alignment = ctr
            else:
                cell.font = black; cell.alignment = lft

    # Group by category for one insert-pass per category
    by_cat = defaultdict(list)
    for r in new_rows:
        by_cat[r["category"].strip().lower()].append(r)

    total_added = 0
    for cat_key, rows in by_cat.items():
        # Locate the TOTAL row for this category (col D = "TOTAL", col A = category)
        total_row = None
        for ws_row in ws.iter_rows():
            if (str(ws_row[3].value or "").strip().upper() == "TOTAL" and
                    str(ws_row[0].value or "").strip().lower() == cat_key):
                total_row = ws_row[0].row
                break

        insert_at = total_row if total_row else ws.max_row + 1
        ws.insert_rows(insert_at, amount=len(rows))

        for offset, r in enumerate(rows):
            write_row(insert_at + offset, r)

        total_added += len(rows)
        log.info("  '%s': %d rows inserted at Excel row %d",
                 cat_key, len(rows), insert_at)

    wb.save(save_path)
    log.info("Excel saved ✓  (total new rows: %d)", total_added)


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPING UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def sanitize(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip(" .")


def make_abs(href: str) -> str:
    if not href:
        return ""
    return href if href.startswith("http") else urljoin(BASE_URL + "/", href.lstrip("/"))


def fname_from_url(url: str) -> str:
    raw = urlparse(url).path.split("/")[-1].split("?")[0]
    raw = re.sub(r'[<>:"/\\|?*]', '_', raw).strip()
    if not raw:
        return ""
    if not raw.lower().endswith(".pdf"):
        raw += ".pdf"
    if raw.lower() in GENERIC_PDF_NAMES:
        return ""
    return raw


def nav(page, url: str, label: str = "") -> bool:
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        page.wait_for_timeout(1_500)
        return True
    except PWTimeout:
        log.warning("  ⚠ Timeout%s: %s", f" ({label})" if label else "", url)
        return False


def get_depts(page) -> list:
    for sel in ["ul#dept_list_content li p a",
                "ul#dept_list_content li a",
                "#dept_container a[href*='dep_id']",
                "a[href*='dep_id']"]:
        try:
            items = page.eval_on_selector_all(sel,
                """els => els
                    .filter(a => a.getAttribute('href') && a.innerText.trim())
                    .map(a => ({name: a.innerText.trim(),
                                href: a.getAttribute('href')}))""")
            if items:
                log.info("    Dept selector '%s' → %d", sel, len(items))
                return items
        except Exception:
            continue
    return []


def get_year_links(page, current_url: str) -> list:
    entries = []
    seen    = {current_url}
    for sel in ["a[href*='year=']", "div.archives a",
                "div.d-flex.flex-wrap a[href*='dep_id']",
                ".archive a", "a[href*='dep_id'][href*='year']"]:
        try:
            items = page.eval_on_selector_all(sel,
                """els => els
                    .filter(a => a.getAttribute('href') && a.innerText.trim())
                    .map(a => ({year: a.innerText.trim(),
                                href: a.getAttribute('href')}))""")
            for it in items:
                ah = make_abs(it["href"])
                yr = it["year"].strip()
                if ah not in seen and yr:
                    seen.add(ah)
                    entries.append({"year": yr, "href": ah})
        except Exception:
            continue
    if not entries:
        entries = [{"year": "", "href": current_url}]
    return entries


def get_pdf_urls(page) -> list:
    found = {}
    for sel in ["a[href$='.pdf']", "a[href*='.pdf']",
                "a[target='_blank'][href*='cms.tn']",
                "table a[href*='.pdf']", "tr td a",
                "div.event-info a", "div.event-card a",
                "p.event-detail a", "div.go-list a",
                "ul li a[href*='.pdf']"]:
        try:
            for u in page.eval_on_selector_all(sel,
                    "els=>els.map(a=>a.href)"
                    ".filter(h=>h&&h.toLowerCase().includes('.pdf'))"):
                found[u] = True
        except Exception:
            continue
    try:
        for u in page.eval_on_selector_all("a",
                "els=>els.map(a=>a.href)"
                ".filter(h=>h&&h.toLowerCase().endsWith('.pdf'))"):
            found[u] = True
    except Exception:
        pass

    valid = []
    for u in found:
        parsed = urlparse(u)
        netloc = parsed.netloc.lower()
        if netloc and "tn.gov.in" not in netloc and "cms.tn" not in netloc:
            continue
        if "tamilthaivazhthusong" in u.lower():
            continue
        valid.append(u)
    return valid


def dl_pdf(url: str, dest: Path, session: requests.Session) -> bool:
    created_parent = False
    parent_dir = dest.parent
    try:
        r = session.get(url, timeout=DL_TIMEOUT, stream=True)
        r.raise_for_status()
        if "html" in r.headers.get("Content-Type", ""):
            log.warning("      ⚠ Got HTML not PDF – skip: %s", url)
            return False
        if not parent_dir.exists():
            parent_dir.mkdir(parents=True, exist_ok=True)
            created_parent = True
        with open(dest, "wb") as fh:
            for chunk in r.iter_content(16_384):
                fh.write(chunk)
        log.info("      ✓ %s (%d KB)", dest.name, dest.stat().st_size // 1024)
        return True
    except Exception as exc:
        log.error("      ✗ %s → %s", url, exc)
        if dest.exists():
            dest.unlink(missing_ok=True)
        if created_parent:
            try:
                parent_dir.rmdir()
            except OSError:
                pass
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  HANDLE ONE PDF  (skip / download / log to Excel)
# ══════════════════════════════════════════════════════════════════════════════

def handle_pdf(pdf_url, category, sub_category,
               dest_folder, excel_index, new_rows, session, stats):
    fname    = fname_from_url(pdf_url)
    if not fname:
        log.info("      ↷ SKIP   unnamed or invalid PDF URL: %s", pdf_url)
        stats["skipped"] += 1
        return

    flow     = fname.lower()
    stripped = re.sub(r'^\d{4}[-_]', '', flow)

    # ── SKIP if already in Excel ─────────────────────────────────────────
    if flow in excel_index or stripped in excel_index:
        log.info("      ↷ SKIP   %s", fname)
        stats["skipped"] += 1
        return

    # ── NEW — download it ─────────────────────────────────────────────────
    log.info("      ↓ NEW    %s", fname)
    time.sleep(DELAY)
    dest = dest_folder / fname
    ok   = dl_pdf(pdf_url, dest, session)

    if ok:
        stats["downloaded"] += 1
        excel_index.add(flow)        # prevent duplicate in same run
        new_rows.append({
            "category":     category,
            "sub_category": sub_category,
            "file_name":    fname,
            "type":         "PDF",
            "pdf_count":    1,
            "word_count":   0,
        })
    else:
        stats["failed"] += 1


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def run():
    log.info("=" * 70)
    log.info("  TN GOVERNMENT PORTAL – EXCEL-ONLY PDF DOWNLOADER")
    log.info("  Started : %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("  Source  : %s", MASTER_EXCEL)
    log.info("  Target  : %s", OUTPUT_EXCEL)
    log.info("=" * 70)

    # Use existing TN.xlsx if it exists (for incremental runs), otherwise start from MASTER_EXCEL
    read_path = Path(OUTPUT_EXCEL) if Path(OUTPUT_EXCEL).exists() else Path(MASTER_EXCEL)
    
    if not read_path.exists():
        log.error("Excel file NOT found at: %s", read_path)
        return

    # ── 1. Build skip-index from Excel ────────────────────────────────────
    excel_index = build_excel_index(str(read_path))

    # ── 2. Output root = "New_Downloads" folder next to the Excel file ────────
    # We use a timestamped folder to keep each run's "New" files separate as requested
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dl_root = read_path.parent / f"New_Downloads_{timestamp}"
    dl_root.mkdir(parents=True, exist_ok=True)
    log.info("  NEW PDFs will be saved to: %s", dl_root)
    log.info("")

    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer": BASE_URL,
    })

    stats    = {"downloaded": 0, "skipped": 0, "failed": 0}
    new_rows = []

    # ── 3. Scrape all categories ──────────────────────────────────────────
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx     = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        for cat in CATEGORIES:
            cat_name  = cat["name"]
            cat_url   = cat["url"]
            has_depts = cat["has_depts"]
            has_years = cat["has_years"]

            # Download folder:  Downloads/<Category>/
            cat_folder = dl_root / sanitize(cat_name)

            log.info("")
            log.info("▓" * 70)
            log.info("  CATEGORY : %s", cat_name)
            log.info("  URL      : %s", cat_url)
            log.info("▓" * 70)

            if not nav(page, cat_url, cat_name):
                log.error("  Cannot load category page — skipping.")
                continue

            # ─────────────────────────────────────────────────────────────
            # PATH A: Categories with Departments  (GO / Policy / Rules / Circulars)
            # Flow: dept list  →  dept page  →  year pages  →  PDFs
            # ─────────────────────────────────────────────────────────────
            if has_depts:
                depts = get_depts(page)
                if not depts:
                    log.error("  No departments found!")
                    try:
                        log.info("  HTML snippet:\n%s", page.content()[:2000])
                    except Exception:
                        pass
                    continue

                log.info("  Departments: %d", len(depts))

                for d_idx, dept in enumerate(depts, 1):
                    dept_name   = sanitize(dept["name"])
                    dept_href   = make_abs(dept["href"])
                    dept_folder = cat_folder / dept_name

                    log.info("")
                    log.info("  ─" * 35)
                    log.info("  [%d/%d] %s", d_idx, len(depts), dept_name)
                    log.info("  ─" * 35)

                    if not nav(page, dept_href, dept_name):
                        continue

                    if has_years:
                        years = get_year_links(page, dept_href)
                        log.info("    Years: %d  %s",
                                 len(years),
                                 [y["year"] for y in years[:8]])

                        first = True
                        for yr in years:
                            yr_url   = yr["href"]
                            yr_folder = dept_folder

                            if not first:
                                if not nav(page, yr_url, dept_name):
                                    continue
                            first = False

                            pdf_urls = get_pdf_urls(page)
                            log.info("    ▶ %-20s  PDFs found: %d",
                                     sanitize(yr["year"]), len(pdf_urls))

                            for pdf_url in pdf_urls:
                                handle_pdf(
                                    pdf_url, cat_name, dept_name,
                                    yr_folder, excel_index,
                                    new_rows, session, stats,
                                )
                    else:
                        # No year level — scrape PDFs directly from dept page
                        pdf_urls = get_pdf_urls(page)
                        log.info("    PDFs found: %d", len(pdf_urls))
                        for pdf_url in pdf_urls:
                            handle_pdf(
                                pdf_url, cat_name, dept_name,
                                dept_folder, excel_index,
                                new_rows, session, stats,
                            )

            # ─────────────────────────────────────────────────────────────
            # PATH B: No Departments  (Acts & Ordinances)
            # Flow: main page  →  year pages  →  PDFs
            # ─────────────────────────────────────────────────────────────
            else:
                # Detect active year label from the page if possible
                active_year = ""
                try:
                    yr_txt = page.eval_on_selector(
                        "h4.event-name, label.text-right, "
                        "div.text-right, span.active",
                        "el => el ? el.innerText.trim() : ''",
                    )
                    if yr_txt and re.match(r'\d{4}', yr_txt):
                        active_year = sanitize(yr_txt)
                except Exception:
                    pass

                year_entries = []
                seen = {cat_url}
                for sel in ["a[href*='year=']", "div.archives a",
                            "div.d-flex.flex-wrap a",
                            f"a[href*='{cat_url.split('/')[-1].split('?')[0]}']"]:
                    try:
                        items = page.eval_on_selector_all(sel,
                            """els => els
                                .filter(a => a.getAttribute('href')
                                          && a.innerText.trim())
                                .map(a => ({year: a.innerText.trim(),
                                            href: a.getAttribute('href')}))""")
                        for it in items:
                            au = make_abs(it["href"])
                            yr = it["year"].strip()
                            if au not in seen and yr:
                                seen.add(au)
                                year_entries.append(
                                    {"year": sanitize(yr), "url": au})
                    except Exception:
                        continue

                if not year_entries:
                    year_entries = [{"year": active_year, "url": cat_url}]

                log.info("  Year pages: %d  %s",
                         len(year_entries),
                         [y["year"] or "(default)" for y in year_entries[:10]])

                first = True
                for yr in year_entries:
                    yr_url   = yr["url"]
                    yr_folder = cat_folder

                    if not first:
                        if not nav(page, yr_url, cat_name):
                            continue
                    first = False

                    pdf_urls = get_pdf_urls(page)
                    log.info("  ▶ %-12s  PDFs found: %d", yr["year"] or "default", len(pdf_urls))

                    for pdf_url in pdf_urls:
                        handle_pdf(
                            pdf_url, cat_name, yr["year"] or cat_name,
                            yr_folder, excel_index,
                            new_rows, session, stats,
                        )

        browser.close()

    # ── 4. Write new rows into Excel ──────────────────────────────────────
    if new_rows:
        log.info("")
        try:
            append_to_excel(str(read_path), new_rows, save_path=OUTPUT_EXCEL)
            log.info("Updated Excel saved to: %s", OUTPUT_EXCEL)
        except Exception as exc:
            log.error("Excel write failed: %s", exc)
            log.error("New rows NOT saved to Excel (listed below):")
            for r in new_rows:
                log.error("  %s", r)
    else:
        log.info("")
        log.info("No new PDFs — Excel is already up to date.")

    # ── 5. Summary ────────────────────────────────────────────────────────
    log.info("")
    log.info("=" * 70)
    log.info("  FINISHED  –  %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 70)
    log.info("  ✓  Downloaded      : %d  new PDFs", stats["downloaded"])
    log.info("  ↷  Skipped         : %d  (filename found in Excel)",
             stats["skipped"])
    log.info("  ✗  Failed          : %d  (see log for details)", stats["failed"])
    log.info("  📄 Excel rows added: %d", len(new_rows))
    log.info("  📁 New PDFs in     : %s", dl_root)
    if new_rows:
        log.info("  📄 Updated Excel    : %s", OUTPUT_EXCEL)
    log.info("=" * 70)


if __name__ == "__main__":
    run()