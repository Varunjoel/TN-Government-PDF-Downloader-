"""
Kerala Finance Department – Incremental PDF Downloader
======================================================
Website : https://finance.kerala.gov.in
 
FEATURES:
  • Reads existing Excel index (Kerala_GOs_Document_Index.xlsx) to skip
    already-downloaded PDFs — comparison is case-insensitive, ignores
    .pdf extension, spaces, and special characters
  • Downloads only NEW PDFs via direct HTTP (fast) with Selenium fallback
  • Appends new rows to the Excel with EXACTLY the same format as the
    original (alternating row colours, per-sheet header colour, borders,
    fonts, column widths, row heights)
  • Updates the S.No column and PDF Count column automatically
  • Saves updated Excel as  Kerala_GOs_Document_Index_updated.xlsx
 
FOLDER STRUCTURE PRODUCED:
  Kerala GOs/
    Government order/<Category>/file.pdf
    Budgets/file.pdf
    Circulars/file.pdf
    Notifications/file.pdf
    Reports/file.pdf
 
SETUP:
  pip install selenium webdriver-manager requests openpyxl pandas
 
USAGE (run from the script folder):
  python Kerala_Downloader.py
"""
 
# ════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  ← only section you ever need to edit
# ════════════════════════════════════════════════════════════════════════
 
EXCEL_INPUT_PATH  = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\Kerala_GOs_Document_Index.xlsx"   # source Excel
EXCEL_OUTPUT_PATH = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\Kerala_GOs_Document_Index_updated.xlsx"  # result
DOWNLOAD_ROOT     = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\Kerala2 GOs"                        # PDF root folder
 
DELAY       = 0.3   # seconds between HTTP requests
DL_TIMEOUT  = 45    # seconds for a single PDF download
PAGE_TIMEOUT = 30_000  # ms for browser page loads
MAX_RETRIES  = 3    # retries per PDF
 
# ════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ════════════════════════════════════════════════════════════════════════
 
import re
import time
import shutil
import logging
import unicodedata
import requests
import urllib3
import pandas as pd
from pathlib import Path
from datetime import datetime
from copy import copy
 
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
 
try:
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    ChromeDriverManager = None
 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
 
# ════════════════════════════════════════════════════════════════════════
#  LOGGING
# ════════════════════════════════════════════════════════════════════════
 
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("kerala_downloader.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)
 
# ════════════════════════════════════════════════════════════════════════
#  SITE / SHEET CONFIGURATION
# ════════════════════════════════════════════════════════════════════════
 
BASE_URL = "https://finance.kerala.gov.in"
 
SECTIONS = [
    {
        "sheet":  "Government Orders",
        "folder": "Government order",
        "url":    f"{BASE_URL}/gos.jsp",
        "is_go":  True,
    },
    {
        "sheet":  "Budgets",
        "folder": "Budgets",
        "url":    f"{BASE_URL}/bdgtDcs.jsp",
        "is_go":  False,
    },
    {
        "sheet":  "Circulars",
        "folder": "Circulars",
        "url":    f"{BASE_URL}/circlr.jsp",
        "is_go":  False,
    },
    {
        "sheet":  "Reports",
        "folder": "Reports",
        "url":    f"{BASE_URL}/rptDocs.jsp",
        "is_go":  False,
    },
    {
        "sheet":  "Notifications",
        "folder": "Notifications",
        "url":    f"{BASE_URL}/ntfctn.jsp",
        "is_go":  False,
    },
]
 
# All 40+ GO categories from <select> on gos.jsp
GO_CATEGORIES = [
    ("BDS",         "Bill Discounting System"),
    ("CSFC",        "Central Finance Commission"),
    ("CMDRS",       "CMs Debt Relief Scheme"),
    ("CMDRF",       "CMs Distress Relief Fund"),
    ("FINPOWER",    "Delegation of Financial Powers"),
    ("EGOV",        "E-Governance"),
    ("EST",         "Establishment"),
    ("EXPCNTRL",    "Expenditure Control"),
    ("FBS",         "Family Benefit Scheme"),
    ("ADVICE",      "Financial Advisory to Administrative Departments"),
    ("FINPLAN",     "Financial Planning for Projects"),
    ("GENINSTRUT",  "General Instructions"),
    ("GIAM",        "Grant In-Aid Monitoring"),
    ("GIS",         "Group Insurance Scheme"),
    ("GPAI",        "Group Personal Accident Insurance Scheme"),
    ("HLTHINSURCE", "Health Insurance MEDISEP"),
    ("HBA",         "House Building Advance"),
    ("I-NT",        "Inspection Non-Technical"),
    ("FININSP",     "Inspection Technical and FIMS"),
    ("IFMS",        "Integrated Financial Management System"),
    ("INTAUDIT",    "Internal Audit"),
    ("LACADF",      "Legislative Assembly Constituency Asset Development Fund"),
    ("LAA",         "Loans and Advances"),
    ("MRS",         "Medical Reimbursement Scheme"),
    ("NPS",         "National Pension System"),
    ("PRC",         "Pay Revision Commission"),
    ("PENSION",     "Pension"),
    ("PRISM",       "PRISM Pensioner Information System"),
    ("PF",          "Provident Fund"),
    ("PERC",        "Public Expenditure Review Committee"),
    ("PUFM",        "Public Undertakings Financial Monitoring"),
    ("PUB",         "Publications"),
    ("REVMON",      "Revenue Monitoring"),
    ("SCORE",       "SCORE State Confidential Reporting"),
    ("SERVRULE",    "Service Rules"),
    ("SWSP",        "Social Welfare Security Pension"),
    ("SPARK",       "SPARK"),
    ("MLASDF",      "Special Development Fund For MLA"),
    ("BUDGET",      "State Budget"),
    ("SFC",         "State Finance Commission"),
    ("SLI",         "State Life Insurance"),
    ("VEELS",       "VEELS Vehicle Management"),
    ("WG",          "Working Group"),
]
 
# Budget year dropdown range (website: 2014–2026)
BUDGET_YEARS = [str(y) for y in range(2014, 2027)]
 
# ════════════════════════════════════════════════════════════════════════
#  EXCEL FORMAT CONSTANTS  (matched exactly from the original file)
# ════════════════════════════════════════════════════════════════════════
 
# Per-sheet header background colour (ARGB, no alpha prefix in openpyxl)
HEADER_COLORS = {
    "Budgets":           "1A5276",   # dark blue
    "Circulars":         "1E8449",   # dark green
    "Government Orders": "922B21",   # dark red
    "Notifications":     "6C3483",   # dark purple
    "Reports":           "784212",   # dark brown
}
 
# Main folder name written into column B for each sheet
MAIN_FOLDER_NAMES = {
    "Budgets":           "Budgets",
    "Circulars":         "Circulars",
    "Government Orders": "Government order",
    "Notifications":     "Notifications",
    "Reports":           "Reports",
}
 
# Alternating row fills
ROW_FILL_LIGHT = PatternFill("solid", fgColor="EBF5FB")   # light blue (odd rows)
ROW_FILL_NONE  = PatternFill("solid", fgColor="FFFFFF")   # white (even rows)
 
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
 
DATA_FONT   = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
 
COL_WIDTHS = {"A": 6, "B": 28, "C": 38, "D": 60, "E": 12}
HEADER_ROW_HEIGHT = 30
 
# ════════════════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════════════
 
def normalize_name(name: str) -> str:
    """
    Lowercase, strip .pdf, collapse whitespace, remove non-alphanumeric,
    normalize unicode.  Used for duplicate detection.
    """
    if not name:
        return ""
    name = str(name).strip()
    name = unicodedata.normalize("NFKD", name)
    name = name.encode("ascii", "ignore").decode("ascii")
    name = name.lower()
    name = re.sub(r"\.pdf$", "", name)
    name = re.sub(r"[^a-z0-9]+", "", name)
    return name
 
 
def sanitize_filename(name: str) -> str:
    """Remove characters not allowed in file names; truncate to 200 chars."""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(name).strip())
    name = re.sub(r"_+", "_", name).strip("_")
    return name[:200] or "document"
 
 
def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path
 
 
# ════════════════════════════════════════════════════════════════════════
#  EXCEL INDEX LOADING
# ════════════════════════════════════════════════════════════════════════
 
def load_excel_index(excel_path: str) -> dict:
    """
    Returns {sheet_name: set_of_normalized_pdf_names}.
    Reads column D (PDF File Name) from every sheet.
    """
    log.info("Loading Excel index: %s", excel_path)
    xf  = pd.ExcelFile(excel_path)
    idx = {}
 
    for sec in SECTIONS:
        sheet = sec["sheet"]
        if sheet not in xf.sheet_names:
            log.warning("  Sheet '%s' not in Excel — treated as empty", sheet)
            idx[sheet] = set()
            continue
 
        df = pd.read_excel(excel_path, sheet_name=sheet, header=0, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
 
        known = set()
        col = "PDF File Name"
        if col in df.columns:
            for v in df[col].dropna():
                n = normalize_name(v)
                if n:
                    known.add(n)
 
        idx[sheet] = known
        log.info("  %-22s : %d known PDFs", sheet, len(known))
 
    return idx
 
 
# ════════════════════════════════════════════════════════════════════════
#  HTTP SESSION
# ════════════════════════════════════════════════════════════════════════
 
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer":        BASE_URL,
        "Accept":         "text/html,application/xhtml+xml,*/*",
        "Accept-Language":"en-US,en;q=0.9",
    })
    try:
        s.get(BASE_URL, timeout=15, verify=False)
    except Exception:
        pass
    return s
 
 
def make_browser() -> webdriver.Chrome:
    """Create a Selenium Chrome driver for page automation."""
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1280,800")
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
 
    service = None
    if ChromeDriverManager is not None:
        try:
            service = Service(ChromeDriverManager().install())
        except Exception:
            service = None
 
    if service is None:
        service = Service()
 
    return webdriver.Chrome(service=service, options=chrome_options)
 
 
def _selenium_wait_for(driver, by, value, timeout_ms=None):
    timeout_s = (timeout_ms or PAGE_TIMEOUT) / 1000
    return WebDriverWait(driver, timeout_s).until(
        EC.presence_of_element_located((by, value))
    )
 
 
# ════════════════════════════════════════════════════════════════════════
#  DOCUMENT LISTING  (fast HTTP, no Selenium)
# ════════════════════════════════════════════════════════════════════════
 
def _parse_docs_from_html(html: str) -> list:
    """
    Extract {doc_id, title, date, view_url} from a search-result HTML page.
    The site uses  dwldDoc('dId')  pattern for every document link.
    """
    docs = []
    seen = set()
 
    row_re  = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
    doc_re  = re.compile(r"(?:dwldDoc|selDoc)\(['\"]([^'\"]+)['\"]\)")
    date_re = re.compile(r"(\d{2}[-/]\d{2}[-/]\d{4})")
 
    for row_m in row_re.finditer(html):
        row_html = row_m.group(1)
        for doc_id in doc_re.findall(row_html):
            doc_id = doc_id.strip()
            if not doc_id or doc_id in seen:
                continue
            seen.add(doc_id)
 
            text  = re.sub(r"<[^>]+>", " ", row_html)
            text  = re.sub(r"\s+", " ", text).strip()
            date_m = date_re.search(text)
            date  = date_m.group(1) if date_m else ""
            # Extract title by removing the date and collapsing whitespace
            if date_m:
                # Remove both DD-MM-YYYY and YYYY-MM-DD (hidden) patterns
                clean_title = re.sub(r"\d{2}[-/]\d{2}[-/]\d{4}", " ", text)
                clean_title = re.sub(r"\d{4}-\d{2}-\d{2}", " ", clean_title)
                title = re.sub(r"\s+", " ", clean_title).strip(" -·|\t")
            else:
                title = text
 
            docs.append({
                "doc_id":   doc_id,
                "title":    title,
                "date":     date,
                "view_url": f"{BASE_URL}/includeWeb/fileViewer.jsp?dId={doc_id}",
            })
 
    return docs
 
 
def _parse_category_options(html: str) -> list:
    """
    Parse ONLY document category options, ignoring year/month/other selects.
    Looks for a select element that contains options like 'BDS', 'Circular', etc.
    """
    # First search for known category select names
    select_pattern = re.compile(
        r'<select[^>]*(?:id|name)=["\'](?:optCategory|optType|category)["\'][^>]*>(.*?)</select>',
        re.IGNORECASE | re.DOTALL
    )
    select_match = select_pattern.search(html)
    if not select_match:
        # fallback: find any select that has more than 1 option and likely contains document types
        all_selects = re.findall(r'<select[^>]*>(.*?)</select>', html, re.DOTALL)
        for opts_html in all_selects:
            if re.search(r'(BDS|Circular|Notification|Budget|Report|PRC|PENSION)', opts_html, re.IGNORECASE):
                select_match = re.search(r'<select[^>]*>(.*?)</select>', opts_html, re.DOTALL)
                break
 
    if not select_match:
        return []
 
    options_html = select_match.group(1)
    opt_re = re.compile(r'<option\s+value=["\']([^"\']+)["\'][^>]*>([^<]+)</option>', re.IGNORECASE)
    options = []
    skip_labels = {"select", "choose", "all", "year", "month", "category", "--", ""}
    for value, label in opt_re.findall(options_html):
        value = value.strip()
        label = label.strip()
        # keep only options that look like document categories (value not numeric, not "0", length reasonable)
        if not value or value == "0" or label.lower() in skip_labels:
            continue
        if re.match(r'^\d+$', value) and len(value) <= 4:   # likely a year
            continue
        options.append((value, label))
    return options
 
 
def _fetch_all_pages(
    session: requests.Session,
    search_url: str,
    base_params: dict,
) -> list:
    """Paginate through all pages for a given search endpoint."""
    all_docs = []
    page_no  = 1
 
    while True:
        params = {**base_params, "pageNo": str(page_no)}
        html   = ""
        try:
            r = session.post(search_url, data=params, timeout=20, verify=False)
            html = r.text if r.status_code == 200 else ""
            if not html:
                r2 = session.get(search_url, params=params, timeout=20, verify=False)
                html = r2.text if r2.status_code == 200 else ""
        except Exception as exc:
            log.warning("    HTTP error (page=%d): %s", page_no, exc)
            break
 
        docs = _parse_docs_from_html(html)
        if not docs:
            break
 
        all_docs.extend(docs)
        log.debug("    page=%d  docs_on_page=%d", page_no, len(docs))
 
        # Stop if no next-page reference
        if not re.search(rf"pageNo={page_no + 1}", html, re.IGNORECASE):
            if not re.search(r"[Nn]ext|>>", html):
                break
 
        page_no += 1
        if page_no > 500:
            break
        time.sleep(0.1)
 
    return all_docs
 
 
def fetch_go_section(session: requests.Session) -> list:
    """Fetch all GO docs across all categories."""
    log.info("  Fetching GOs (%d categories) via HTTP …", len(GO_CATEGORIES))
    all_docs = []
 
    for code, label in GO_CATEGORIES:
        docs = _fetch_all_pages(
            session,
            f"{BASE_URL}/includeWeb/goSearch.jsp",
            {"category": code, "srchType": "Ct"},
        )
        for d in docs:
            d["category_code"] = code
            d["category_name"] = label
        all_docs.extend(docs)
        log.info("    %-45s : %d docs", label[:45], len(docs))
        time.sleep(0.1)
 
    log.info("  GOs total: %d", len(all_docs))
    return all_docs
 
 
def fetch_budget_section(session: requests.Session, driver) -> list:
    """
    Budget page has a year dropdown (2014–2026).
    Uses Selenium to select each year and collect doc IDs.
    """
    log.info("  Fetching Budgets (year dropdown 2014–2026) …")
    all_docs = []
    seen     = set()
 
    try:
        driver.get(f"{BASE_URL}/bdgtDcs.jsp")
        _selenium_wait_for(driver, By.ID, "optYear")
        time.sleep(1.5)
    except Exception as exc:
        log.warning("  Could not load budget page: %s", exc)
        return all_docs
 
    try:
        select = Select(driver.find_element(By.ID, "optYear"))
    except Exception as exc:
        log.warning("  Budget page missing year selector: %s", exc)
        return all_docs
 
    for year in BUDGET_YEARS:
        try:
            select.select_by_value(year)
            time.sleep(1.5)
            html = driver.page_source
            docs = _parse_docs_from_html(html)
            imported = 0
            for d in docs:
                if d["doc_id"] not in seen:
                    seen.add(d["doc_id"])
                    d["category_name"] = f"Budget {year}"
                    all_docs.append(d)
                    imported += 1
            log.info("    Year %s : %d docs (new %d)", year, len(docs), imported)
        except Exception as exc:
            log.warning("    Year %s failed: %s", year, exc)
        time.sleep(0.2)
 
    log.info("  Budgets total: %d", len(all_docs))
    return all_docs
 
 
def _fetch_notifications_selenium(driver, url):
    """Specific scraper for Notifications page."""
    log.info("  Using Notifications-specific Selenium scraper")
    all_docs = []
    seen = set()
 
    try:
        driver.get(url)
        _selenium_wait_for(driver, By.TAG_NAME, "table")
        time.sleep(1.5)
    except Exception as exc:
        log.warning("  Failed to load Notifications page: %s", exc)
        return []
 
    page_num = 1
    while page_num <= 7:
        html = driver.page_source
        docs = _parse_docs_from_html(html)
        new_count = 0
        for d in docs:
            if d["doc_id"] not in seen:
                seen.add(d["doc_id"])
                all_docs.append(d)
                new_count += 1
        log.info("    Page %d: Found %d docs (%d new)", page_num, len(docs), new_count)
 
        clicked = False
        next_selectors = [
            "#dynamic-table_next:not(.disabled) a",
            "//li[@id='dynamic-table_next' and not(contains(@class,'disabled'))]//a",
            "//a[normalize-space()='Next']",
            "//li[contains(@class,'next')]//a",
        ]
        for sel in next_selectors:
            try:
                if sel.startswith("//"):
                    elements = driver.find_elements(By.XPATH, sel)
                else:
                    elements = driver.find_elements(By.CSS_SELECTOR, sel)
                
                for el in elements:
                    if el.is_displayed() and el.is_enabled():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", el)
                        time.sleep(2.0)
                        clicked = True
                        break
                if clicked:
                    break
            except Exception:
                continue
 
        if not clicked:
            break
        
        page_num += 1
 
    log.info("  Found %d documents via Selenium", len(all_docs))
    return all_docs
 
 
def _fetch_generic_selenium(driver, url, name):
    """Generic scraper for any section – load and extract available docs."""
    log.info("  Using generic Selenium scraper for %s", name)
    all_docs = []
    seen = set()
 
    try:
        driver.get(url)
        time.sleep(3.0)  # Allow JS rendering
    except Exception as exc:
        log.warning("  Selenium load failed for %s: %s", name, exc)
        return []
 
    # Attempt to parse docs from initial page load
    html = driver.page_source
    docs = _parse_docs_from_html(html)
   
    for d in docs:
        if d["doc_id"] not in seen:
            seen.add(d["doc_id"])
            all_docs.append(d)
 
    log.info("  Found %d documents via Selenium on initial page load", len(all_docs))
   
    # If no docs found on initial load, log warning and return
    if len(all_docs) == 0:
        log.warning("  No documents found on %s page - may require manual category selection", name)
   
    return all_docs
 
 
def fetch_section_docs(
    session: requests.Session,
    url: str,
    name: str,
    driver=None,
) -> list:
    """
    Fetch documents for a section (Circulars, Notifications, Reports).
    Uses Selenium scraper for Circulars, Reports, and Notifications (robust handling).
    """
    log.info("  Fetching %s …", name)
 
    # ---- Selenium-based scraping for Circulars/Reports/Notifications ----
    # (Direct HTTP category search doesn't work reliably for these sections)
    if driver is not None:
        if name == "Notifications":
            return _fetch_notifications_selenium(driver, url)
        elif name in ("Circulars", "Reports"):
            return _fetch_generic_selenium(driver, url, name)
 
    # ---- Fallback: HTTP attempt (unlikely to succeed but try) ----
    try:
        r0 = session.get(url, timeout=20, verify=False)
        html0 = r0.text
    except Exception as exc:
        log.warning("  Could not load %s: %s", url, exc)
        return []
 
    categories = _parse_category_options(html0)
    search_url = f"{BASE_URL}/includeWeb/docSearch.jsp"
 
    if categories:
        log.info("  Found %d categories via HTTP", len(categories))
        all_docs = []
        for code, label in categories:
            try:
                docs = _fetch_all_pages(session, search_url,
                                        {"category": code, "srchType": "Ct"})
                if docs:
                    for d in docs:
                        d["category_name"] = label
                    all_docs.extend(docs)
                    log.info("    %-45s : %d docs", label[:45], len(docs))
            except Exception as exc:
                log.debug("    Category %s failed: %s", code, exc)
            time.sleep(0.1)
        if all_docs:
            return all_docs
        else:
            log.info("  HTTP gave 0 docs – would need Selenium")
 
    # ---- Fallback: Selenium generic scraper ----
    if driver is None:
        log.warning("  No Selenium driver provided – returning empty")
        return []
    return _fetch_generic_selenium(driver, url, name)
 
 
# ════════════════════════════════════════════════════════════════════════
#  PDF DOWNLOAD
# ════════════════════════════════════════════════════════════════════════
 
def _extract_filename_from_response(r: requests.Response, doc_id: str) -> str:
    """Derive a PDF filename from Content-Disposition or URL."""
    cd    = r.headers.get("Content-Disposition", "")
    m     = re.search(r'filename[^;=\n]*=["\']?([^"\';\n]+)', cd)
    fname = m.group(1).strip().strip('"\'') if m else ""
    if not fname:
        from urllib.parse import urlparse
        fname = urlparse(r.url).path.split("/")[-1].split("?")[0]
    if not fname or not fname.lower().endswith(".pdf"):
        fname = f"{doc_id}.pdf"
    return sanitize_filename(fname)
 
 
def _is_valid_pdf(path: Path) -> bool:
    try:
        with open(path, "rb") as fh:
            return fh.read(4) == b"%PDF"
    except Exception:
        return False
 
 
def _try_direct_download(
    doc_id: str,
    dest_folder: Path,
    session: requests.Session,
) -> tuple:
    """
    Try several direct HTTP URL patterns.
    Returns (filename, True) on success, ("", False) on failure.
    """
    patterns = [
        f"{BASE_URL}/includeWeb/getPdf.jsp?dId={doc_id}",
        f"{BASE_URL}/includeWeb/downloadFile.jsp?dId={doc_id}",
        f"{BASE_URL}/includeWeb/fileViewer.jsp?dId={doc_id}",
        f"{BASE_URL}/downloadPdf?dId={doc_id}",
    ]
    for url in patterns:
        for attempt in range(MAX_RETRIES):
            try:
                r = session.get(url, timeout=DL_TIMEOUT, stream=True,
                                allow_redirects=True, verify=False)
                if r.status_code != 200:
                    break
                ct = r.headers.get("Content-Type", "")
                if "html" in ct and "pdf" not in ct:
                    break   # not a PDF response
 
                fname = _extract_filename_from_response(r, doc_id)
                ensure_dir(dest_folder)
                dest  = dest_folder / fname
 
                with open(dest, "wb") as fh:
                    for chunk in r.iter_content(16_384):
                        fh.write(chunk)
 
                if not _is_valid_pdf(dest):
                    dest.unlink(missing_ok=True)
                    break
 
                kb = dest.stat().st_size // 1024
                log.info("    ✓ DOWNLOADED  %s  (%d KB)", fname, kb)
                return fname, True
 
            except Exception as exc:
                log.debug("    attempt %d failed for %s: %s", attempt + 1, url, exc)
                time.sleep(1)
 
    return "", False
 
 
def _selenium_download(
    driver,
    doc_id: str,
    view_url: str,
    dest_folder: Path,
    session: requests.Session,
) -> tuple:
    """
    Selenium fallback: open the fileViewer page, locate PDF sources,
    then download with requests.
    """
    pdf_urls = []
 
    try:
        driver.get(view_url)
        time.sleep(1.0)
    except Exception:
        pass
 
    try:
        current_url = driver.current_url
        if current_url.lower().endswith(".pdf"):
            pdf_urls.append(current_url)
    except Exception:
        pass
 
    for tag in ("iframe", "embed", "object"):
        try:
            elements = driver.find_elements(By.TAG_NAME, tag)
            for el in elements:
                src = el.get_attribute("src") or el.get_attribute("data")
                if src and "pdf" in src.lower():
                    pdf_urls.append(src)
        except Exception:
            pass
 
    try:
        html = driver.page_source
        pdf_urls.extend(re.findall(r'https?://[^"\']+\.pdf', html, re.IGNORECASE))
    except Exception:
        html = ""
 
    if not pdf_urls:
        try:
            r = session.get(view_url, timeout=20, verify=False)
            if r.status_code == 200:
                pdf_urls.extend(re.findall(r'href=["\']([^"\']+\.pdf)["\']', r.text, re.IGNORECASE))
                pdf_urls.extend(re.findall(r'src=["\']([^"\']+\.pdf)["\']', r.text, re.IGNORECASE))
        except Exception:
            pass
 
    pdf_urls = [u for u in dict.fromkeys(pdf_urls) if u]
    for pu in pdf_urls:
        if not pu.startswith("http"):
            from urllib.parse import urljoin
            pu = urljoin(BASE_URL + "/", pu.lstrip("/"))
        try:
            r = session.get(pu, timeout=DL_TIMEOUT, stream=True,
                            allow_redirects=True, verify=False)
            if r.status_code != 200:
                continue
            fname = _extract_filename_from_response(r, doc_id)
            ensure_dir(dest_folder)
            dest  = dest_folder / fname
            with open(dest, "wb") as fh:
                for chunk in r.iter_content(16_384):
                    fh.write(chunk)
            if not _is_valid_pdf(dest):
                dest.unlink(missing_ok=True)
                continue
            kb = dest.stat().st_size // 1024
            log.info("    ✓ DOWNLOADED (Selenium) %s  (%d KB)", fname, kb)
            return fname, True
        except Exception as exc:
            log.debug("    Selenium url download failed: %s", exc)
 
    return "", False
 
 
def download_pdf(
    doc: dict,
    dest_folder: Path,
    session: requests.Session,
    driver,
) -> tuple:
    """
    Master download function.
    Returns (filename, success_bool).
    """
    doc_id   = doc["doc_id"]
    view_url = doc["view_url"]
 
    # Check physical disk first
    existing = list(dest_folder.glob(f"*{doc_id}*"))
    if existing:
        log.info("    ↷ SKIPPED (file exists on disk) %s", existing[0].name)
        return existing[0].name, True
 
    fname, ok = _try_direct_download(doc_id, dest_folder, session)
    if ok:
        return fname, True
 
    # Selenium fallback
    log.info("    ⚡ Falling back to Selenium for %s", doc_id)
    fname, ok = _selenium_download(driver, doc_id, view_url, dest_folder, session)
    if ok:
        return fname, True
 
    log.warning("    ✗ FAILED  %s", doc_id)
    return "", False
 
 
# ════════════════════════════════════════════════════════════════════════
#  SKIP CHECK
# ════════════════════════════════════════════════════════════════════════
 
def is_known(doc: dict, known_set: set) -> bool:
    """
    Return True if this document is already indexed in the Excel.
    Normalises the doc_id and any title-derived guess.
    """
    did = normalize_name(doc.get("doc_id", ""))
    if did and did in known_set:
        return True
    # doc_id.pdf pattern
    if normalize_name(doc.get("doc_id", "") + ".pdf") in known_set:
        return True
    # Sometimes the title contains the GO number used as file name
    title_n = normalize_name(doc.get("title", ""))
    if title_n and title_n in known_set:
        return True
    return False
 
 
# ════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT  — preserves EXACT original format
# ════════════════════════════════════════════════════════════════════════
 
def _make_header_fill(sheet_name: str) -> PatternFill:
    color = HEADER_COLORS.get(sheet_name, "1A5276")
    return PatternFill("solid", fgColor=color)
 
 
def _row_fill(row_number: int) -> PatternFill:
    """
    Alternating: data rows start at row 2.
    Odd data rows (2,4,6…) → light blue; even (3,5,7…) → white.
    Same as original: row index within data (1-based) odd → light.
    """
    data_index = row_number - 1   # row 2 → data_index 1
    return ROW_FILL_LIGHT if (data_index % 2 == 1) else ROW_FILL_NONE
 
 
def append_rows_to_excel(
    input_path: str,
    output_path: str,
    new_rows_by_sheet: dict,
):
    """
    Copy the original Excel to output_path then append new rows to each
    sheet, preserving the EXACT format of the original file.
    """
    shutil.copy2(input_path, output_path)
 
    if not any(new_rows_by_sheet.values()):
        log.info("No new rows — Excel copied unchanged.")
        return
 
    wb = load_workbook(output_path)
 
    for sec in SECTIONS:
        sheet_name = sec["sheet"]
        rows       = new_rows_by_sheet.get(sheet_name, [])
        if not rows:
            continue
 
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Write header
            headers = ["S.No", "Main Folder", "Sub Folder / Department",
                       "PDF File Name", "PDF Count"]
            hfill = _make_header_fill(sheet_name)
            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill      = hfill
                cell.font      = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border    = THIN_BORDER
            ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
            for col_letter, width in COL_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
        else:
            ws = wb[sheet_name]
 
        # Find last data row
        last_row = ws.max_row
        while last_row > 1 and ws.cell(row=last_row, column=1).value is None:
            last_row -= 1
 
        # Compute next S.No
        last_sno = ws.cell(row=last_row, column=1).value
        try:
            next_sno = int(last_sno) + 1
        except (TypeError, ValueError):
            next_sno = last_row  # fallback
 
        main_folder = MAIN_FOLDER_NAMES.get(sheet_name, sheet_name)
 
        for i, row_data in enumerate(rows):
            rn       = last_row + 1 + i
            sno      = next_sno + i
            fill     = _row_fill(rn)
            sub_folder = row_data.get("category_name") or "[Directly in Folder]"
 
            values = [
                sno,
                main_folder,
                sub_folder,
                row_data.get("filename", ""),
                1,
            ]
            alignments = [CENTER_ALIGN, LEFT_ALIGN, LEFT_ALIGN, LEFT_ALIGN, CENTER_ALIGN]
 
            for ci, (val, aln) in enumerate(zip(values, alignments), start=1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.fill      = fill
                cell.font      = DATA_FONT
                cell.alignment = aln
                cell.border    = THIN_BORDER
 
        log.info("  Sheet %-22s : +%d rows appended (from row %d)",
                 sheet_name, len(rows), last_row + 1)
 
    wb.save(output_path)
    log.info("Updated Excel saved → %s", output_path)
 
 
# ════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════
 
def run():
    log.info("=" * 70)
    log.info("  KERALA FINANCE – INCREMENTAL PDF DOWNLOADER")
    log.info("  %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 70)
 
    excel_in  = Path(EXCEL_INPUT_PATH)
    excel_out = Path(EXCEL_OUTPUT_PATH)
    dl_root   = Path(DOWNLOAD_ROOT)
 
    if not excel_in.exists():
        log.error("Excel not found: %s", excel_in)
        return
 
    ensure_dir(dl_root)
    log.info("  Input  Excel : %s", excel_in)
    log.info("  Output Excel : %s", excel_out)
    log.info("  PDF folder   : %s", dl_root)
 
    # Load existing index
    indexes = load_excel_index(str(excel_in))
 
    session = make_session()
 
    stats = {"downloaded": 0, "skipped": 0, "failed": 0}
    new_rows_by_sheet = {sec["sheet"]: [] for sec in SECTIONS}
 
    try:
        driver = make_browser()
    except Exception as exc:
        log.error("Could not start Selenium browser: %s", exc)
        return
 
    try:
        for sec in SECTIONS:
            sheet_name  = sec["sheet"]
            folder_name = sec["folder"]
            url         = sec["url"]
            is_go       = sec.get("is_go", False)
            known       = indexes[sheet_name]
            sec_dir     = ensure_dir(dl_root / folder_name)
 
            log.info("")
            log.info("▓" * 70)
            log.info("  SECTION : %s  (known: %d)", folder_name, len(known))
            log.info("▓" * 70)
 
            # ── Fetch document listing ─────────────────────────────────
            if is_go:
                docs = fetch_go_section(session)
            elif folder_name == "Budgets":
                docs = fetch_budget_section(session, driver)
            else:
                # Notifications / Circulars / Reports
                docs = fetch_section_docs(session, url, folder_name, driver)
 
            if not docs:
                log.warning("  No documents found for %s", folder_name)
                continue
 
            new_count = sum(1 for d in docs if not is_known(d, known))
            log.info("  Site total : %d  |  New : %d  |  Skip : %d",
                     len(docs), new_count, len(docs) - new_count)
 
            # ── Process each document ──────────────────────────────────
            for doc in docs:
                doc_id   = doc["doc_id"]
                cat_name = doc.get("category_name", "")
 
                if is_known(doc, known):
                    log.debug("  ↷ SKIPPED (in Excel)  %s", doc_id)
                    stats["skipped"] += 1
                    continue
 
                # Destination folder
                if is_go and cat_name:
                    dest_folder = ensure_dir(sec_dir / sanitize_filename(cat_name))
                else:
                    dest_folder = sec_dir
 
                log.info("  ↓ [%s]  %s",
                         (cat_name or folder_name)[:20],
                         doc.get("title", "")[:60])
 
                time.sleep(DELAY)
 
                fname, ok = download_pdf(doc, dest_folder, session, driver)
 
                if ok and fname:
                    stats["downloaded"] += 1
                    # Mark as known so we don't re-download in same run
                    known.add(normalize_name(fname))
                    known.add(normalize_name(doc_id))
                    new_rows_by_sheet[sheet_name].append({
                        "filename":      fname,
                        "doc_id":        doc_id,
                        "title":         doc.get("title", ""),
                        "date":          doc.get("date", ""),
                        "view_url":      doc.get("view_url", ""),
                        "category_name": cat_name,
                    })
                else:
                    stats["failed"] += 1
 
            section_new = len(new_rows_by_sheet[sheet_name])
            log.info("  ── %s done : ↓%d  ↷%d  ✗%d",
                     folder_name, section_new,
                     stats["skipped"], stats["failed"])
    finally:
        try:
            driver.quit()
        except Exception:
            pass
 
    # ── Write updated Excel ────────────────────────────────────────────
    total_new = sum(len(v) for v in new_rows_by_sheet.values())
    log.info("")
    log.info("Writing Excel: %d new rows …", total_new)
    try:
        append_rows_to_excel(
            str(excel_in),
            str(excel_out),
            new_rows_by_sheet,
        )
    except Exception as exc:
        log.error("Excel write error: %s", exc)
 
    # ── Summary ────────────────────────────────────────────────────────
    log.info("")
    log.info("=" * 70)
    log.info("  DONE  %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 70)
    log.info("  ✓ Downloaded  : %d", stats["downloaded"])
    log.info("  ↷ Skipped    : %d  (already in Excel / on disk)", stats["skipped"])
    log.info("  ✗ Failed     : %d", stats["failed"])
    log.info("  📄 Excel rows : +%d", total_new)
    log.info("")
    log.info("  PDF output   : %s/", dl_root)
    log.info("  Excel output : %s", excel_out)
    log.info("=" * 70)
 
 
if __name__ == "__main__":
    run()
 