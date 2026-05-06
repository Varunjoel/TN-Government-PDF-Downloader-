"""
Mydoc Folder Scanner - Generates Excel report of PDF & Word documents
Usage: python scan_mydoc.py
Requirements: pip install openpyxl
"""
 
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
 
# ─────────────────────────────────────────────
#  ✅ CHANGE THIS PATH TO YOUR MYDOC FOLDER
# ─────────────────────────────────────────────
MYDOC_PATH = r"C:\Users\varun\OneDrive\ドキュメント\Desktop\DEPARTMENT- ORG RTI\TN GOs"
OUTPUT_FILE = r"TN Report.xlsx"
# ─────────────────────────────────────────────
 
PDF_EXTS  = {".pdf"}
WORD_EXTS = {".doc", ".docx"}
 
def get_doc_type(filename):
    ext = Path(filename).suffix.lower()
    if ext in PDF_EXTS:
        return "PDF"
    if ext in WORD_EXTS:
        return "Word"
    return None
 
def scan_folder(folder_path):
    """
    Returns list of dicts:
      { subcategory: str|None, filename: str, doc_type: str }
    Recursively walks all sub-levels.
    """
    records = []
    base = Path(folder_path)
 
    for item in sorted(base.rglob("*")):
        if item.is_file():
            doc_type = get_doc_type(item.name)
            if not doc_type:
                continue
            # Build subcategory path relative to category folder
            rel = item.relative_to(base)
            parts = rel.parts  # e.g. ('SubA', 'deep', 'file.pdf')
            if len(parts) == 1:
                subcat = None   # directly in category folder
            else:
                subcat = str(Path(*parts[:-1]))  # e.g. 'SubA' or 'SubA\deep'
            records.append({
                "subcategory": subcat,
                "filename": item.name,
                "doc_type": doc_type,
            })
    return records
 
# ── Styles ──────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
 
def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)
 
COLOR_MAIN_HDR  = "1F3864"   # dark navy  – column headers
COLOR_CAT_HDR   = "2E75B6"   # blue       – category row
COLOR_SUBCAT    = "D6E4F0"   # light blue – subcategory group
COLOR_TOTAL_CAT = "BDD7EE"   # medium blue– per-category totals
COLOR_GRAND     = "1F3864"   # dark navy  – grand total row
COLOR_ALT_ROW   = "F2F7FC"   # faint blue – alternating data rows
 
def apply_font(cell, bold=False, color="000000", size=10):
    cell.font = Font(name="Arial", bold=bold, color=color, size=size)
 
def apply_fill(cell, hex_color):
    cell.fill = PatternFill("solid", start_color=hex_color, fgColor=hex_color)
 
def apply_border(cell):
    cell.border = thin_border()
 
def center(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
def left(cell):
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
# ── Build Excel ──────────────────────────────
def build_excel(scan_results, output_path):
    wb = Workbook()
 
    # ── Sheet 1: Detail ──────────────────────
    ws = wb.active
    ws.title = "Document Details"
    ws.freeze_panes = "A2"
 
    # Column widths
    col_widths = [30, 35, 45, 12, 10, 10]
    cols = ["A","B","C","D","E","F"]
    for c, w in zip(cols, col_widths):
        ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 30
 
    # Header row
    headers = ["Category", "Sub-Category", "File Name", "Type", "PDF Count", "Word Count"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        apply_fill(cell, COLOR_MAIN_HDR)
        apply_font(cell, bold=True, color="FFFFFF", size=11)
        center(cell)
        apply_border(cell)
 
    current_row = 2
    summary_rows = []   # for Sheet 2
 
    grand_pdf = 0
    grand_word = 0
    grand_total = 0
 
    alt = False  # alternating row flag
 
    for cat_name, records in sorted(scan_results.items()):
        cat_pdf  = sum(1 for r in records if r["doc_type"] == "PDF")
        cat_word = sum(1 for r in records if r["doc_type"] == "Word")
        cat_total = cat_pdf + cat_word
 
        # ── Category header row ───────────────
        cat_row = current_row
        for col_idx in range(1, 7):
            cell = ws.cell(row=cat_row, column=col_idx)
            apply_fill(cell, COLOR_CAT_HDR)
            apply_font(cell, bold=True, color="FFFFFF", size=11)
            apply_border(cell)
            center(cell)
        ws.cell(row=cat_row, column=1, value=cat_name)
        ws.cell(row=cat_row, column=2, value="(All sub-folders)")
        ws.cell(row=cat_row, column=3, value=f"Total files in category: {cat_total}")
        ws.cell(row=cat_row, column=5, value=cat_pdf)
        ws.cell(row=cat_row, column=6, value=cat_word)
        left(ws.cell(row=cat_row, column=1))
        current_row += 1
 
        # Group records by subcategory
        subcat_groups = {}
        for r in records:
            key = r["subcategory"] or "(Root)"
            subcat_groups.setdefault(key, []).append(r)
 
        for subcat_name, sub_records in sorted(subcat_groups.items()):
            sub_pdf  = sum(1 for r in sub_records if r["doc_type"] == "PDF")
            sub_word = sum(1 for r in sub_records if r["doc_type"] == "Word")
 
            # ── Subcategory label row ─────────
            sc_row = current_row
            for col_idx in range(1, 7):
                cell = ws.cell(row=sc_row, column=col_idx)
                apply_fill(cell, COLOR_SUBCAT)
                apply_font(cell, bold=True, color="1F3864", size=10)
                apply_border(cell)
                center(cell)
            ws.cell(row=sc_row, column=2, value=subcat_name)
            ws.cell(row=sc_row, column=3, value=f"Files: {len(sub_records)}")
            ws.cell(row=sc_row, column=5, value=sub_pdf)
            ws.cell(row=sc_row, column=6, value=sub_word)
            left(ws.cell(row=sc_row, column=2))
            current_row += 1
 
            # ── Individual file rows ──────────
            for r in sorted(sub_records, key=lambda x: x["filename"]):
                fill_hex = COLOR_ALT_ROW if alt else "FFFFFF"
                alt = not alt
                for col_idx in range(1, 7):
                    cell = ws.cell(row=current_row, column=col_idx)
                    apply_fill(cell, fill_hex)
                    apply_font(cell, size=10)
                    apply_border(cell)
                    left(cell)
 
                ws.cell(row=current_row, column=1, value=cat_name)
                ws.cell(row=current_row, column=2, value=subcat_name)
                ws.cell(row=current_row, column=3, value=r["filename"])
                type_cell = ws.cell(row=current_row, column=4, value=r["doc_type"])
                center(type_cell)
                if r["doc_type"] == "PDF":
                    apply_fill(type_cell, "FFE0E0")
                    apply_font(type_cell, bold=True, color="C00000", size=10)
                else:
                    apply_fill(type_cell, "E0EFF9")
                    apply_font(type_cell, bold=True, color="2E75B6", size=10)
                ws.cell(row=current_row, column=5, value=1 if r["doc_type"] == "PDF"  else 0)
                ws.cell(row=current_row, column=6, value=1 if r["doc_type"] == "Word" else 0)
                current_row += 1
 
        # ── Per-category total row ────────────
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx)
            apply_fill(cell, COLOR_TOTAL_CAT)
            apply_font(cell, bold=True, color="1F3864", size=10)
            apply_border(cell)
            center(cell)
        ws.cell(row=current_row, column=1, value=f"TOTAL — {cat_name}")
        ws.cell(row=current_row, column=4, value="TOTAL")
        ws.cell(row=current_row, column=5, value=cat_pdf)
        ws.cell(row=current_row, column=6, value=cat_word)
        left(ws.cell(row=current_row, column=1))
        ws.cell(row=current_row, column=3, value=f"PDF: {cat_pdf}  |  Word: {cat_word}  |  Total: {cat_total}")
        current_row += 2   # blank gap between categories
 
        grand_pdf  += cat_pdf
        grand_word += cat_word
        grand_total += cat_total
        summary_rows.append((cat_name, cat_pdf, cat_word, cat_total))
 
    # ── Grand Total row ───────────────────────
    for col_idx in range(1, 7):
        cell = ws.cell(row=current_row, column=col_idx)
        apply_fill(cell, COLOR_GRAND)
        apply_font(cell, bold=True, color="FFFFFF", size=12)
        apply_border(cell)
        center(cell)
    ws.cell(row=current_row, column=1, value="GRAND TOTAL (All 20 Categories)")
    ws.cell(row=current_row, column=3, value=f"PDF: {grand_pdf}  |  Word: {grand_word}  |  Total: {grand_total}")
    ws.cell(row=current_row, column=4, value="ALL")
    ws.cell(row=current_row, column=5, value=grand_pdf)
    ws.cell(row=current_row, column=6, value=grand_word)
    ws.row_dimensions[current_row].height = 25
 
    # ── Sheet 2: Summary ─────────────────────
    ws2 = wb.create_sheet("Category Summary")
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.row_dimensions[1].height = 28
 
    hdrs2 = ["Category Name", "PDF Count", "Word Count", "Total Count"]
    for ci, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        apply_fill(cell, COLOR_MAIN_HDR)
        apply_font(cell, bold=True, color="FFFFFF", size=11)
        center(cell)
        apply_border(cell)
 
    for ri, (cat_name, cpdf, cword, ctotal) in enumerate(summary_rows, 2):
        fill_hex = COLOR_ALT_ROW if ri % 2 == 0 else "FFFFFF"
        data = [cat_name, cpdf, cword, ctotal]
        for ci, val in enumerate(data, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            apply_fill(cell, fill_hex)
            apply_font(cell, size=10)
            apply_border(cell)
            if ci == 1:
                left(cell)
            else:
                center(cell)
 
    # Summary grand total
    gt_row = len(summary_rows) + 2
    ws2.row_dimensions[gt_row].height = 22
    for ci, val in enumerate(["GRAND TOTAL", grand_pdf, grand_word, grand_total], 1):
        cell = ws2.cell(row=gt_row, column=ci, value=val)
        apply_fill(cell, COLOR_GRAND)
        apply_font(cell, bold=True, color="FFFFFF", size=11)
        apply_border(cell)
        center(cell)
 
    wb.save(output_path)
    return grand_pdf, grand_word, grand_total
 
 
# ── Main ─────────────────────────────────────
def main():
    print(f"\n📂 Scanning: {MYDOC_PATH}")
    if not os.path.isdir(MYDOC_PATH):
        print(f"❌ ERROR: Folder not found → {MYDOC_PATH}")
        print("   Please update MYDOC_PATH at the top of this script.")
        return
 
    # Get all direct sub-folders (categories)
    all_items = sorted(Path(MYDOC_PATH).iterdir())
    categories = [p for p in all_items if p.is_dir()]
 
    if not categories:
        print("❌ No category folders found inside Mydoc.")
        return
 
    print(f"✅ Found {len(categories)} category folders.\n")
 
    scan_results = {}
    for cat_path in categories:
        cat_name = cat_path.name
        records  = scan_folder(cat_path)
        scan_results[cat_name] = records
        pdf_c  = sum(1 for r in records if r["doc_type"] == "PDF")
        word_c = sum(1 for r in records if r["doc_type"] == "Word")
        print(f"   📁 {cat_name:45s} PDF: {pdf_c:4d}  Word: {word_c:4d}")
 
    print(f"\n⚙️  Building Excel report …")
    g_pdf, g_word, g_total = build_excel(scan_results, OUTPUT_FILE)
 
    print(f"\n✅ Report saved → {OUTPUT_FILE}")
    print(f"\n{'─'*50}")
    print(f"   Grand Total PDF  : {g_pdf}")
    print(f"   Grand Total Word : {g_word}")
    print(f"   Grand Total ALL  : {g_total}")
    print(f"{'─'*50}\n")
 
 
if __name__ == "__main__":
    main()
 